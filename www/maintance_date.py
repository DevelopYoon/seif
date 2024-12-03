import pandas as pd
import shutil
import requests
from PIL import Image
from io import BytesIO
import os
import tempfile
import pymysql
import json
from django.conf import settings
from sqlalchemy import create_engine
from openpyxl.drawing.image import Image as OpenpyxlImage
import openpyxl

def get_sqlalchemy_engine():
    db_info = settings.DATABASES['default']
    db_url = f"mysql+pymysql://{db_info['USER']}:{db_info['PASSWORD']}@{db_info['HOST']}:{db_info['PORT']}/{db_info['NAME']}"
    return create_engine(db_url)

def handle_images(wb, merged_df, place, check_date):
    image_urls = merged_df['picture'].dropna().tolist()
    if not image_urls:
        return

    images_per_sheet = 6
    total_sheets = (len(image_urls) // images_per_sheet) + 1
    
    for sheet_index in range(total_sheets):
        base_sheet_name = f'사진첩_{place}_{check_date}'
        
        if base_sheet_name not in wb.sheetnames:
            # 새 시트 추가
            template_photos_sheet = wb['사진첩']
            new_sheet = wb.copy_worksheet(template_photos_sheet)
            new_sheet.title = base_sheet_name
        else:
            new_sheet = wb[base_sheet_name]

        start_index = sheet_index * images_per_sheet
        end_index = start_index + images_per_sheet
        images_to_insert = image_urls[start_index:end_index]

        for i, url in enumerate(images_to_insert):
            cell = ['A6', 'F6', 'A14', 'F14', 'A22', 'F22'][i]
            img_data = download_image(url)
            if img_data:
                try:
                    # 임시 파일 생성
                    with tempfile.NamedTemporaryFile(dir='/home/safeit/seif/media/tmp', delete=False, suffix=".jpg") as temp_file:
                        temp_file.write(img_data.getvalue())
                        temp_file_path = temp_file.name
                    

                    # 엑셀 파일에 이미지 삽입
                    img = OpenpyxlImage(temp_file_path)
                    img.width = 280
                    img.height = 224
                    ws = wb[base_sheet_name]
                    img.anchor = cell
                    ws.add_image(img)
                    
                    # 파일 삭제
                    # os.remove(temp_file_path)
                except Exception as e:
                    print(f"이미지 처리 중 오류 발생: {e}")
            else:
                print(f"이미지 다운로드 또는 삽입 실패: {url}")
                
        # check_date를 B5 셀에 삽입
        new_sheet['B5'] = check_date

def download_image(url):
    if pd.isna(url) or not url.startswith(('http://', 'https://')):
        print(f"유효하지 않은 URL: {url}")
        return None
    try:
        response = requests.get(url)
        response.raise_for_status()
        return BytesIO(response.content)
    except requests.exceptions.HTTPError as err:
        print(f"HTTP 오류 발생: {err} - URL: {url}")
    except Exception as err:
        print(f"기타 오류 발생: {err} - URL: {url}")
    return None


def parse_and_insert_detail(wb, sheet_path, detail, start_row, sheet_name):
    try:
        # JSON 데이터 로드
        details = json.loads(detail)
        
        # JSON 데이터를 DataFrame으로 변환
        df_detail = pd.DataFrame(details)
        
        # 열 이름 설정
        if not df_detail.empty:
            # 유니코드 이스케이프 시퀀스를 사람이 읽을 수 있는 문자열로 변환
            df_detail.columns = [col.encode('utf-8').decode('utf-8') for col in df_detail.columns]

            # 인덱스 재설정
            df_detail.index += 1  # A 열에 번호 넣기
            df_detail.index.name = None

            # 열 위치 정의
            columns = {
                '번호': 'A',
                '키': 'B',
                'H': 'H',
                'I': 'I'
            }

            # Excel 파일에 데이터 삽입
            # wb = openpyxl.load_workbook(sheet_path)

            # 시트 이름이 이미 존재하는지 확인
            if sheet_name in wb.sheetnames:
                new_sheet = wb[sheet_name]  # 이미 존재하는 시트를 사용
            else:
                # 새로운 시트 생성
                new_sheet = wb.create_sheet(sheet_name)

            # 데이터 삽입
            for idx, row in df_detail.iterrows():
                current_row = start_row + idx
                for col_idx, (key, value) in enumerate(row.items()):
                    if pd.notna(value):  # 값이 NaN이 아닌 경우에만 처리
                        print(f"Inserting data at row {current_row}")
                        new_sheet[f"{columns['번호']}{current_row}"] = idx  # 번호 삽입
                        if idx > 0 and (idx - 1) < len(df_detail.index):
                            new_sheet[f'A{current_row}'] = df_detail.index[idx - 1]
                        else:
                            print(f"Index {idx - 1} is out of bounds for df_detail")
                        new_sheet[f"{columns['키']}{current_row}"] = key    # 키 삽입

                        # 'value'에 따라 H 또는 I 열에 'V'를 입력
                        if value == 1:
                            new_sheet[f"{columns['H']}{current_row}"] = 'V'
                        elif value == 0:
                            new_sheet[f"{columns['I']}{current_row}"] = 'V'                                    
        
        # else:
        #     print("No data to insert.")
        
    except json.JSONDecodeError as e:
        print(f"detail 파싱 오류: {e}")
    except Exception as e:
        print(f"기타 오류: {e}")



def process_excel_file(processType, facilityNum, place, checkDate, isLast = True):
    engine = get_sqlalchemy_engine()  # DB 엔진 생성
        
    query2 = """
        SELECT co_Num, place, date, writer, detail
        FROM tb_maintenanceList L
        WHERE co_Num = %s
          AND place = %s
          AND date = %s
        ORDER BY place ASC
        """
 
    query3 = """
        SELECT P.facilityNum, P.place, P.picture, P.maintenance, P.checkDate
        FROM tb_maintenancePic P
        JOIN tb_maintenanceList L 
          ON P.facilityNum = L.co_Num
          AND P.checkDate = L.date
          AND P.place = L.place
        WHERE P.facilityNum = %s
          AND P.checkDate = %s
          AND P.place = %s
        ORDER BY P.checkDate ASC
        """      

    try:
        
        with engine.connect() as conn:
            df2 = pd.read_sql(query2, conn, params=(facilityNum, place, checkDate))
            df3 = pd.read_sql(query3, conn, params=(facilityNum, checkDate, place))
        
        merged_df = pd.merge(df2, df3, left_on=['co_Num', 'date', 'place'], right_on=['facilityNum', 'checkDate', 'place'], how='outer')

        excel_path = '/home/safeit/seif/media/유지관리표.xlsx'
        fileDir = f'/home/safeit/seif/media/exel/{facilityNum}'
        
        if processType==1:
            output_path = f'{fileDir}/유지관리표_{place}_{checkDate}.xlsx'
        elif processType==2:
            output_path = f'{fileDir}/유지관리표_{place}.xlsx'
        elif processType==3:
            output_path = f'{fileDir}/유지관리표_총합본.xlsx'
            
        output_path = output_path.replace(" ", "_")

        if not os.path.exists(fileDir):
            os.makedirs(fileDir, exist_ok=True)

        else:
            print(f"디렉토리가 이미 존재합니다: {fileDir}")

        if os.path.isfile(output_path):
            print(f"파일이 존재합니다: {output_path}")
        else:
            shutil.copy(excel_path, output_path)


        template_checklist_sheet = '체크리스트'
        
        # 파일에 시트가 적어도 하나 있는지 확인
        wb = openpyxl.load_workbook(output_path)

        if wb is None:
            print("엑셀 파일을 열 수 없습니다.")
            return None

        # 그룹화된 데이터프레임
        grouped_df = merged_df.groupby(['co_Num', 'date', 'writer', 'place'])
        
        for name, group in grouped_df:
            place = group['place'].iloc[0]
            date = group['date'].iloc[0]
            new_sheet_name = f'체크리스트_{place}_{date}'[:31]

            # 시트 이름이 이미 존재하는 경우 처리
            if new_sheet_name in wb.sheetnames:
                wb.remove(wb[new_sheet_name])
                
            # 새 시트 추가
            template_sheet = wb[template_checklist_sheet]
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = new_sheet_name

            # 기본 데이터 삽입
            group = group.reset_index(drop=True)  # 인덱스 재설정
            detail = group['detail'].iloc[0]  # 그룹의 첫 번째 detail만 사용
            
            ws = wb[new_sheet_name]
            ws['E6'] = group['writer'].iloc[0]
            ws['H6'] = group['place'].iloc[0]
            # ws['B5'] = group['checkDate'].iloc[0]
            ws['B5'] = date
            
  
            image_urls = merged_df['picture'].dropna().tolist()
            if image_urls:
                handle_images(wb, merged_df, place, date)
            else:
                print("이미지가 없습니다. 이미지 처리 생략")  
            
            parse_and_insert_detail(wb, output_path, detail, 7, new_sheet_name)    


        if isLast:
            del wb['사진첩']
            del wb['체크리스트']
        else :
            print("마지막 아님!!!!!!!!!!!!!!!!")


        # 엑셀 파일 저장
        wb.save(output_path)
        
        wb.close()

        file_path = '/'.join(output_path.split('/')[4:])

        file_url = f"http://211.237.0.230:10123/{file_path}"
        return file_url
    
    except Exception as e:
        print(f"오류 발생: {e}")