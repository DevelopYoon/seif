import subprocess
import pandas as pd
import shutil
import requests
from PIL import Image
from io import BytesIO
import os
import tempfile
import pymysql
import json
import re
from django.conf import settings
from sqlalchemy import create_engine
from openpyxl.drawing.image import Image as OpenpyxlImage
import openpyxl
from datetime import datetime

# SQLAlchemy 엔진 생성
def get_sqlalchemy_engine():
    db_info = settings.DATABASES['default']
    db_url = f"mysql+pymysql://{db_info['USER']}:{db_info['PASSWORD']}@{db_info['HOST']}:{db_info['PORT']}/{db_info['NAME']}"
    return create_engine(db_url)

def extract_number_dot(text):
    """
    문자열에서 '숫자.숫자' 형태의 부분만 추출합니다.
    """
    matches = re.findall(r'(\d+\.\d+)', text)
    return matches if matches else []

def fill_symbols(sheet, number_dots):
    """
    숫자.숫자 값에 따라 특정 셀에 기호를 삽입합니다.
    """
    symbol = '■'
    cell_mapping = {
        '1.1': 'O5', '1.2': 'X5', '1.3': 'AH5', '1.4': 'AQ5', '1.5': 'P6', '1.6': 'X6', '1.7': 'AF6',
        '2.1': 'O7', '2.2': 'U7', '2.3': 'AC7', '2.4': 'AN7', '2.5': 'AV7', '2.6': 'O8', '2.7': 'W8',
        '2.8': 'AH8', '2.9': 'AQ8', '2.10': 'O9', '2.11': 'Y9',
        '3.1': 'O10', '3.2': 'Y10', '3.3': 'AI10', '3.4': 'AS10', '3.5': 'O11', '3.6': 'Y11', '3.7': 'AH11',
        '3.8': 'AQ11',
        '4.1': 'O12', '4.2': 'S12', '4.3': 'W12', '4.4': 'AE12', '4.5': 'AM12', '4.6': 'AW12', '4.7': 'BB12',
        '5.1': 'O13', '5.2': 'V13', '5.3': 'AA13', '5.4': 'AF13', '5.5': 'AQ13', '5.6': 'O14', '5.7': 'AA14',
        '5.8': 'AC14', '5.9': 'AH14', '5.10': 'O15', '5.11': 'AB15',
        '6.1': 'O16', '6.2': 'AC16', '6.3': 'AP16', '6.4': 'AW16', '6.5': 'O17', '6.6': 'W17', '6.7': 'AH17',
        '6.8': 'AS17', '6.9': 'O18', '6.10': 'T18', '6.11': 'AA18'
    }
    
    for number_dot in number_dots:
        if number_dot in cell_mapping:
            cell = cell_mapping[number_dot]
            sheet[cell] = symbol
            
def download_image(url):
    """URL에서 이미지를 다운로드합니다."""
    if pd.isna(url) or not url.startswith(('http://', 'https://')):
        print(f"유효하지 않은 URL: {url}")
        return None
    try:
        response = requests.get(url)
        response.raise_for_status()
        img = Image.open(BytesIO(response.content))
        return BytesIO(response.content)
    except requests.exceptions.HTTPError as err:
        print(f"HTTP 오류 발생: {err} - URL: {url}")
    except Exception as err:
        print(f"기타 오류 발생: {err} - URL: {url}")
    return None

def insert_image(sheet, url, cell):
    """엑셀 시트의 특정 셀에 이미지를 삽입합니다."""
    img_data = download_image(url)

    if img_data:
        try:
            # 임시 파일 생성 (자동 삭제를 위해 delete=True 사용)
            with tempfile.NamedTemporaryFile(dir='media/tmp', delete=False, suffix=".jpg") as temp_file:
                temp_file.write(img_data.getvalue())  # BytesIO 객체의 바이트 내용 기록
                temp_file_path = temp_file.name

            # 엑셀 파일에 이미지 삽입
            img = OpenpyxlImage(temp_file_path)
            img.width = 305
            img.height = 128
            img.anchor = cell
            sheet.add_image(img)

        except Exception as e:
            print(f"이미지 처리 중 오류 발생: {e}")
    else:
        print(f"이미지 다운로드 또는 삽입 실패: {url}")




def queryCoNum(co_Num):
    query = """
            SELECT  *
            FROM tb_problem 
            WHERE co_Num = %s
            ORDER BY firstDate ASC, place ASC, work ASC, bigCause ASC 
            """
    return query        

def queryPlace(co_Num, place):
    query = """
            SELECT  *
            FROM tb_problem
            WHERE co_Num = %s
              AND place = %s
            ORDER BY firstDate ASC, place ASC, work ASC, bigCause ASC 
            """
    return query   

def queryWork(co_Num, place, work):
    query = """
            SELECT *
            FROM tb_problem
            WHERE co_Num = %s
              AND place = %s
              AND work = %s
            ORDER BY firstDate ASC, place ASC, work ASC, bigCause ASC
            """
    return query                       



def createRiskDateExcel(process_type, co_number, place=None, work=None):

    engine = get_sqlalchemy_engine()

    # SQL 쿼리 실행 (특정 테이블의 특정 변수만 선택)

    query = None
    params = None
    if process_type == 1:
        query = queryCoNum(co_number)
        params = (co_number,)
    elif process_type == 2:
        query = queryPlace(co_number, place)
        params = (co_number, place)
    elif process_type == 3:
        query = queryWork(co_number, place, work)
        params = (co_number, place, work)

    try:
        with engine.connect() as conn:
            df = pd.read_sql(query, conn, params=params)       

        query2 = """
                SELECT *
                FROM tb_solution
                WHERE (co_Num, place, work) IN (
                    SELECT co_Num, place, work
                    FROM tb_problem
                ORDER BY 'record' ASC
                )
                """
                
        query3 = """
                    SELECT co_Num, name
                    FROM tb_facility
                    WHERE co_Num = %s
                """
        
        query4 = """
                    SELECT co_Num, username, CEO_permission, safe_permission
                    FROM tb_user
                    WHERE co_Num = %s
                """     
                       
        df2 = pd.read_sql(query2, engine)
        # df와 df2를 병합
        df_merged = pd.merge(df, df2, left_on=['co_Num', 'place', 'work'], 
                            right_on=['co_Num', 'place', 'work'], how='left')
        # print(f"{df_merged}")
        
        df3 = pd.read_sql(query3, engine, params=(co_number,))
        df_merged2 = pd.merge(df_merged, df3, left_on='co_Num', right_on='co_Num', how='left')
        
        df4 = pd.read_sql(query4, engine, params=(co_number,))
        # df_final = pd.merge(df_merged2, df4, left_on='co_Num', right_on='co_Num', how='left')
        # print(f"{df_final}")
        
        # 'midCause'에서 숫자.숫자 부분만 추출
        df_merged['midCause_number_dot'] = df_merged['midCause'].apply(extract_number_dot)

        # 엑셀 파일 열기
        excel_path = "/home/safeit/seif/media/위험성평가.xlsx"
        fileDir = f'/home/safeit/seif/media/exel/{co_number}'
        if process_type==3:
            output_path = f'{fileDir}/위험성평가표_{place}_{work}.xlsx'
        elif process_type==2:
            output_path = f'{fileDir}/위험성평가표_{place}.xlsx'
        elif process_type==1:
            output_path = f'{fileDir}/위험성평가표_총합본.xlsx'

        output_path = output_path.replace(" ", "_")

        if not os.path.exists(fileDir):
            os.makedirs(fileDir, exist_ok=True)
            print(f"디렉토리를 생성했습니다: {fileDir}")
        else:
            print(f"디렉토리가 이미 존재합니다: {fileDir}")

        if os.path.isfile(output_path):
            print(f"파일이 존재합니다: {output_path}")
        else:
            # 복사본 파일 생성
            print(f"파일이 존재하지 않습니다: {output_path}")
            shutil.copy(excel_path, output_path)
        
        wb = openpyxl.load_workbook(output_path)

        # 템플릿 시트 선택
        template_sheet5 = wb['표지']
        template_sheet6 = wb['위험성 결정']
        template_sheet = wb['위험성평가표']
        template_sheet2 = wb['작업단계별 유해위험요인 확인표']
        template_sheet3 = wb['위험성관리 계획서(1)']
        template_sheet4 = wb['위험성관리 계획서 (2)']
        
        current_sheet5 = None
        
        if not df_merged2.empty:
            row = df_merged2.iloc[0]  # 첫 번째 행 선택

            # 새로운 시트 이름 생성
            new_sheet_name5 = f'표지_{row["name"]}'[:31]
            
            # 시트 복사
            current_sheet5 = wb.copy_worksheet(template_sheet5)
            current_sheet5.title = new_sheet_name5
            
            # 데이터 입력
            current_sheet5['F22'] = row['name']
            
            # 현재 날짜 입력
            current_date = datetime.today().strftime('%Y-%m-%d')
            current_sheet5['I27'] = current_date
            
            new_sheet_name6 = f'위험성 결정_{row["name"]}'[:31]
            current_sheet6 = wb.copy_worksheet(template_sheet6)
            current_sheet6.title = new_sheet_name6

        # 시트 저장을 위한 변수
        current_sheet2 = None
        sheet_counter2 = 1

        for i, row in df_merged.iterrows():
            new_sheet_name2 = f'확인표_{row["place"]}_{row["work"]}_{row["firstDate"]}'[:31]
            

            if pd.notna(row['work']) and row['work'].strip():
                if current_sheet2 is None or current_sheet2.title != new_sheet_name2:
                    new_sheet = wb.copy_worksheet(template_sheet2)
                    if new_sheet is None:
                        raise ValueError("새 시트를 생성할 수 없습니다. template_sheet2가 올바르지 않습니다.")
                    new_sheet.title = new_sheet_name2
                    current_sheet2 = new_sheet
                    row_start = 8  # 새 시트에서 데이터 삽입 시작 행 번호 초기화
                    # 페이지 번호 설정
                    current_sheet2['AQ1'] = sheet_counter2
                    sheet_counter2 += 1  # 새로운 시트가 생성될 때마다 시트 카운터 증가
                
                current_sheet2['G3'] = row['place']
                current_sheet2['AG3'] = row['work']
                # 기호 채우기
                number_dots = row['midCause_number_dot']
                fill_symbols(current_sheet2, number_dots)
        
        current_sheet = None
        current_sheet_name = ""
        row_start = 8  # 데이터가 삽입될 시작 행 번호
        sheet_counter = 1  # 시트 카운터 초기화
        page_counter = 1  # 페이지 카운터 초기화
        register_num = 1

        # 데이터의 위치를 기억하기 위한 딕셔너리
        current_data_count = 0

        # 변수들 초기화
        last_place = None
        last_work = None
        last_firstDate = None
        last_record = None
        
        for i, row in df_merged.iterrows():
            base_sheet_name = f'평가표_{row["place"]}_{row["work"]}_{row["firstDate"]}'

            # record, place, work, firstDate 값이 변경될 때 새로운 시트 이름 생성
            if pd.notna(row['work']) and row['work'].strip():
                if (row["place"] == last_place and row["work"] == last_work and row["firstDate"] == last_firstDate and row["record"] != last_record):
                    new_sheet_name = f'평가표_{row["place"]}_{row["work"]}_{row["record"]}'[:31]
                    current_sheet = wb.copy_worksheet(template_sheet)
                    current_sheet.title = new_sheet_name
                    row_start = 8  # 새 시트에서 데이터 삽입 시작 행 번호 초기화
                    current_sheet['V4'] = sheet_counter
                    sheet_counter += 1
                    current_data_count = 0  # 항목 수 초기화
                else:
                    new_sheet_name = f'{base_sheet_name}_{page_counter}'[:31]

                    # 새로운 시트가 필요할 때
                    if current_sheet_name != new_sheet_name or current_data_count >= 6:
                        if current_data_count >= 6:
                            page_counter += 1

                        new_sheet_name = f'{base_sheet_name}_{page_counter}'[:31]
                        current_sheet = wb.copy_worksheet(template_sheet)
                        current_sheet.title = new_sheet_name
                        current_sheet_name = new_sheet_name
                        row_start = 8  # 새 시트에서 데이터 삽입 시작 행 번호 초기화
                        current_sheet['V4'] = sheet_counter
                        sheet_counter += 1
                        current_data_count = 0  # 항목 수 초기화


                # 데이터 삽입
                current_sheet['AF2'] = row['writer_x']
                current_sheet['E3'] = row['place']
                current_sheet['AE3'] = row['work']
                current_sheet['AE4'] = row['peopleList']
                current_sheet['B14'] = row['record'] if not pd.isna(row['record']) else row['firstDate']
                current_sheet[f'E{row_start}'] = row['midCause']

                # midCause에서 첫 번째 숫자 추출
                if row['midCause']:
                    number_matches = re.findall(r'\d+', row['midCause'])
                    first_number = number_matches[0] if number_matches else ''
                else:
                    first_number = ''

                current_sheet[f'B{row_start}'] = first_number
                current_sheet[f'F{row_start}'] = row['detail']
                current_sheet[f'K{row_start}'] = row['law']
                current_sheet[f'Q{row_start}'] = row['dangerSolutionBefore']
                current_sheet[f'Y{row_start}'] = row['frequency']
                current_sheet[f'AA{row_start}'] = row['strength']
                current_sheet[f'AC{row_start}'] = row['riskScore']
                current_sheet[f'AG{row_start}'] = row['dangerSolutionAfter']
                current_sheet[f'AE{row_start}'] = register_num

            # 항목 수 증가
                current_data_count += 1
                row_start += 1
            register_num += 1

            # 현재 row의 값을 저장
            last_place = row["place"]
            last_work = row["work"]
            last_firstDate = row["firstDate"]
            last_record = row["record"]


        # 위험성관리계획1    
        current_sheet3 = None
        current_sheet_name3 = ""
        sheet_counter3 = 0
        row_start3 = 6
        register_num3 = 1
        
        for i, row in df_merged.iterrows():
            new_sheet_name3 = f'계획서1_{row["place"]}_{row["work"]}_{row["firstDate"]}'[:31]

            if pd.notna(row['dangerSolutionAfter']):
                if current_sheet_name3 != new_sheet_name3:
                    current_sheet3 = wb.copy_worksheet(template_sheet3)
                    current_sheet3.title = new_sheet_name3
                    current_sheet_name3 = new_sheet_name3
                    sheet_counter3 += 1
                    row_start3 = 6
                    
                    
                    if not df4.empty:
                        for j in range(len(df4)):
                            if df4.iloc[j]['CEO_permission'] == 'Y':
                                current_sheet3['AT3'] = df4.iloc[j]['username']

                            # safe_permission 값이 'Y'인 경우
                            if df4.iloc[j]['safe_permission'] == 'Y':
                                current_sheet3['AN2'] = df4.iloc[j]['username']
                                current_sheet3['AN3'] = df4.iloc[j]['username']
            
                current_sheet3['Y2'] = sheet_counter3
                current_sheet3['E2'] = row['place']
                current_sheet3['AT2'] = row['writer_x']
                
                current_sheet3[f'A{row_start3}'] = register_num3
                current_sheet3[f'J{row_start3}'] = row['frequency']
                current_sheet3[f'L{row_start3}'] = row['strength']
                current_sheet3[f'N{row_start3}'] = row['riskScore']
                current_sheet3[f'P{row_start3}'] = row['dangerSolutionAfter']
                # current_sheet3[f'AN{row_start3}'] = row['record']
                if i + 1 < len(df_merged):
                    next_row3 = df_merged.iloc[i + 1]
                    current_sheet3[f'AH{row_start3}'] = next_row3['frequency']
                    current_sheet3[f'AJ{row_start3}'] = next_row3['strength']
                    current_sheet3[f'AL{row_start3}'] = next_row3['riskScore']
                    # current_sheet3[f'AR{row_start3}'] = next_row3['record']
                else:
                    # 다음 행이 없을 경우에 대한 처리 (예: 빈 값 설정)
                    current_sheet3[f'AH{row_start3}'] = ""
                    current_sheet3[f'AJ{row_start3}'] = ""
                    current_sheet3[f'AL{row_start3}'] = ""
                    # current_sheet3[f'AR{row_start3}'] = ""
                    
                row_start3 += 1
            register_num3 += 1

    
        # 위험성관리계획2    
        current_sheet4 = None
        current_sheet_name4 = ""
        row_start4 = 6
        before_positions = ['J6', 'J10', 'J15', 'J20']
        after_positions = ['P6', 'P10', 'P15', 'P20']
        image_count = 0
        sheet_counter4 = 0
        register_num4 = 1

        for i, row in df_merged.iterrows():
            new_sheet_name4 = f'계획서2_{row["place"]}_{row["work"]}_{row["firstDate"]}'[:31]

            # 새로운 시트가 필요할 때 시트 생성
            if (row.get("pic_before") and pd.notna(row["pic_before"])) or (row.get("pic_after") and pd.notna(row["pic_after"])):
                if current_sheet_name4 != new_sheet_name4:
                    current_sheet4 = wb.copy_worksheet(template_sheet4)
                    current_sheet4.title = new_sheet_name4
                    current_sheet_name4 = new_sheet_name4
                    sheet_counter4 += 1
                    image_count = 0
                    row_start4 = 6
                    
                    if not df4.empty:
                        for j in range(len(df4)):
                            if df4.iloc[j]['CEO_permission'] == 'Y':
                                current_sheet4['AT3'] = df4.iloc[j]['username']

                            # safe_permission 값이 'Y'인 경우
                            if df4.iloc[j]['safe_permission'] == 'Y':
                                current_sheet4['AN2'] = df4.iloc[j]['username']
                                current_sheet4['AN3'] = df4.iloc[j]['username']

                pic_before_url = row["pic_before"]
                pic_after_url = row["pic_after"]

                current_sheet4['R2'] = sheet_counter4
                current_sheet4['E2'] = row['place']
                current_sheet4['AT2'] = row['writer_x']

            # 현재 row의 'record' 값을 셀에 삽입
            if (row.get("pic_before") and pd.notna(row["pic_before"])) or (row.get("pic_after") and pd.notna(row["pic_after"])):
                current_sheet4[f'AN{row_start4}'] = row['record']
                current_sheet4[f'A{row_start4}'] = register_num4

            # 이미지 삽입
            try:
                if pic_before_url:
                    insert_image(current_sheet4, pic_before_url, before_positions[image_count % 4])
                
                if pic_after_url:
                    insert_image(current_sheet4, pic_after_url, after_positions[image_count % 4])
            except Exception as e:
                    print(f"이미지없음")
                    
            pic_before_url = 'nan'
            pic_after_url = 'nan'

            image_count += 1
            register_num4 += 1
            if row_start4 == 6:
                row_start4 += 4
            else:
                row_start4+=5

            if image_count % 4 == 0 and i + 1 < len(df_merged):
                next_row = df_merged.iloc[i + 1]
                next_pic_before_url = next_row["pic_before"]
                next_pic_after_url = next_row["pic_after"]

                # 다음 행에 이미지가 하나라도 있을 경우에만 새 시트 생성
                if (row["workstep_y"]==next_row["workstep_y"]) and (next_pic_before_url or next_pic_after_url):
                    current_sheet4 = wb.copy_worksheet(template_sheet4)
                    current_sheet4.title = f'{new_sheet_name4}_{sheet_counter4}'
                    sheet_counter4 += 1
                    image_count = 0  # 새 시트에서 이미지 카운트를 초기화
                    row_start4 = 6
                    
                    if not df4.empty:
                        for j in range(len(df4)):
                            if df4.iloc[j]['CEO_permission'] == 'Y':
                                current_sheet4['AT3'] = df4.iloc[j]['username']

                            # safe_permission 값이 'Y'인 경우
                            if df4.iloc[j]['safe_permission'] == 'Y':
                                current_sheet4['AN2'] = df4.iloc[j]['username']
                                current_sheet4['AN3'] = df4.iloc[j]['username']
                    
                
            # 템플릿 시트 삭제
        if '표지' in wb.sheetnames:
            del wb['표지']
        if '위험성 결정' in wb.sheetnames:
            del wb['위험성 결정']
        if '위험성평가표' in wb.sheetnames:
            del wb['위험성평가표']
        if '작업단계별 유해위험요인 확인표' in wb.sheetnames:
            del wb['작업단계별 유해위험요인 확인표']    
        if '위험성관리 계획서(1)' in wb.sheetnames:
            del wb['위험성관리 계획서(1)']       
        if '위험성관리 계획서 (2)' in wb.sheetnames:
            del wb['위험성관리 계획서 (2)']              
        
        wb.save(output_path)
        wb.close()
        
        
        # /home/safeit/seif/ 디렉토리 추출
        file_path = '/'.join(output_path.split('/')[4:7]) + '/'

        # 결과물 파일명 .pdf
        file_name = output_path.split('/')[-1].replace('.xlsx', '.pdf')

        # PDF 변환
        subprocess.run(f"libreoffice --headless --convert-to pdf {output_path}", shell=True)

        # 변환된 PDF 파일의 경로
        pdf_output = '/'.join(output_path.split('/')[:4])+'/'+file_name

        # PDF 파일을 결과 경로로 이동
        if os.path.exists(pdf_output):
            destination = file_path + file_name
            
            # 동일한 이름의 파일이 있으면 삭제 후 이동
            if os.path.exists(destination):
                os.remove(destination)
            
            shutil.move(pdf_output, destination)
        else:
            raise FileNotFoundError(f"PDF 변환에 실패했습니다: {pdf_output}")

        # 엑셀 파일 삭제
        if os.path.exists(output_path):
            os.remove(output_path)
        
        # 파일 URL 반환
        file_url = f"http://211.237.0.230:10123/{file_path + file_name}"
        return file_url
         
    except Exception as e:
        print(f"엑셀 애플리케이션 종료 중 오류 발생: {e}")

